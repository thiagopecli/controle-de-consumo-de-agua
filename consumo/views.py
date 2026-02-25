from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.db.models import Sum, Avg, Max, Min, Count, Q, Case, When, Value, IntegerField
from django.http import HttpResponse
from django.conf import settings
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from datetime import timedelta, datetime
import json
import io
import os
import zipfile
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.utils import ImageReader

from .models import Lote, Hidrometro, Leitura
from .serializers import (
    LoteSerializer, 
    HidrometroSerializer, 
    LeituraSerializer,
    LeituraCreateSerializer
)


# Mapeamento de meses em português do Brasil
MESES_PT_BR = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}


def formatar_mes_ano_ptbr(data):
    """Retorna string formatada 'Mês/Ano' em português do Brasil"""
    mes_nome = MESES_PT_BR[data.month]
    return f"{mes_nome}/{data.year}"


def _obter_caminho_logo_marca_dagua():
    caminhos = [
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.jpeg'),
        os.path.join(settings.BASE_DIR, 'logo.jpeg'),
    ]
    for caminho in caminhos:
        if os.path.exists(caminho):
            return caminho
    return None


def _desenhar_marca_dagua_logo(canvas, doc):
    caminho_logo = _obter_caminho_logo_marca_dagua()
    if not caminho_logo:
        return

    try:
        pagina_largura, pagina_altura = doc.pagesize
        imagem = ImageReader(caminho_logo)
        img_largura, img_altura = imagem.getSize()
        proporcao = img_altura / float(img_largura) if img_largura else 1

        logo_largura = pagina_largura * 0.45
        logo_altura = logo_largura * proporcao

        if logo_altura > pagina_altura * 0.6:
            logo_altura = pagina_altura * 0.6
            logo_largura = logo_altura / proporcao if proporcao else logo_largura

        pos_x = (pagina_largura - logo_largura) / 2
        pos_y = (pagina_altura - logo_altura) / 2

        canvas.saveState()
        if hasattr(canvas, 'setFillAlpha'):
            canvas.setFillAlpha(0.08)
        canvas.drawImage(
            caminho_logo,
            pos_x,
            pos_y,
            width=logo_largura,
            height=logo_altura,
            preserveAspectRatio=True,
            mask='auto'
        )
        canvas.restoreState()
    except Exception:
        return


class LoteViewSet(viewsets.ModelViewSet):
    """API endpoint para gerenciar lotes"""
    queryset = Lote.objects.all()
    serializer_class = LoteSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero', 'endereco']
    ordering_fields = ['numero', 'tipo', 'criado_em']
    ordering = ['numero']
    
    @action(detail=True, methods=['get'])
    def hidrometros(self, request, pk=None):
        """Retorna todos os hidrômetros de um lote"""
        lote = self.get_object()
        hidrometros = lote.hidrometros.filter(ativo=True)
        serializer = HidrometroSerializer(hidrometros, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def consumo_total(self, request, pk=None):
        """Retorna o consumo total de um lote em um período"""
        lote = self.get_object()
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')
        
        if not data_inicio or not data_fim:
            return Response(
                {'error': 'Parâmetros data_inicio e data_fim são obrigatórios'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        leituras = Leitura.objects.filter(
            hidrometro__lote=lote,
            data_leitura__range=[data_inicio, data_fim]
        )
        
        consumo_total = 0
        for hidrometro in lote.hidrometros.filter(ativo=True):
            leituras_h = leituras.filter(hidrometro=hidrometro).order_by('data_leitura')
            if leituras_h.exists():
                primeira = leituras_h.first()
                ultima = leituras_h.last()
                consumo_total += float(ultima.leitura - primeira.leitura)
        
        return Response({
            'lote': lote.numero,
            'periodo': f'{data_inicio} a {data_fim}',
            'consumo_total_m3': consumo_total
        })


class HidrometroViewSet(viewsets.ModelViewSet):
    """API endpoint para gerenciar hidrômetros"""
    queryset = Hidrometro.objects.all()
    serializer_class = HidrometroSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero', 'lote__numero', 'localizacao']
    ordering_fields = ['numero', 'data_instalacao', 'lote__numero']
    ordering = ['numero']
    
    def get_queryset(self):
        queryset = Hidrometro.objects.all()
        lote_id = self.request.query_params.get('lote', None)
        ativo = self.request.query_params.get('ativo', None)
        
        if lote_id:
            queryset = queryset.filter(lote_id=lote_id)
        if ativo is not None:
            queryset = queryset.filter(ativo=ativo.lower() == 'true')
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def leituras_periodo(self, request, pk=None):
        """Retorna leituras de um hidrômetro em um período"""
        hidrometro = self.get_object()
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')
        
        if not data_inicio or not data_fim:
            return Response(
                {'error': 'Parâmetros data_inicio e data_fim são obrigatórios'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        leituras = hidrometro.leituras.filter(
            data_leitura__range=[data_inicio, data_fim]
        ).order_by('data_leitura')
        
        serializer = LeituraSerializer(leituras, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def estatisticas(self, request, pk=None):
        """Retorna estatísticas de consumo de um hidrômetro"""
        hidrometro = self.get_object()
        dias = int(request.query_params.get('dias', 30))
        
        data_inicio = timezone.now() - timedelta(days=dias)
        leituras = hidrometro.leituras.filter(data_leitura__gte=data_inicio).order_by('data_leitura')
        
        if not leituras.exists():
            return Response({'message': 'Sem leituras no período especificado'})
        
        primeira_leitura = leituras.first()
        ultima_leitura = leituras.last()
        consumo_total = float(ultima_leitura.leitura - primeira_leitura.leitura)
        consumo_medio_dia = consumo_total / dias if dias > 0 else 0
        
        return Response({
            'hidrometro': hidrometro.numero,
            'periodo_dias': dias,
            'total_leituras': leituras.count(),
            'consumo_total_m3': consumo_total,
            'consumo_medio_dia_m3': round(consumo_medio_dia, 3),
            'primeira_leitura': primeira_leitura.leitura,
            'ultima_leitura': ultima_leitura.leitura
        })


class LeituraViewSet(viewsets.ModelViewSet):
    """API endpoint para gerenciar leituras"""
    queryset = Leitura.objects.all()
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['hidrometro__numero', 'hidrometro__lote__numero', 'responsavel']
    ordering_fields = ['data_leitura', 'leitura']
    ordering = ['-data_leitura']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return LeituraCreateSerializer
        return LeituraSerializer
    
    def get_queryset(self):
        queryset = Leitura.objects.all()
        hidrometro_id = self.request.query_params.get('hidrometro', None)
        data_inicio = self.request.query_params.get('data_inicio', None)
        data_fim = self.request.query_params.get('data_fim', None)
        periodo = self.request.query_params.get('periodo', None)
        
        if hidrometro_id:
            queryset = queryset.filter(hidrometro_id=hidrometro_id)
        if data_inicio:
            queryset = queryset.filter(data_leitura__gte=data_inicio)
        if data_fim:
            queryset = queryset.filter(data_leitura__lte=data_fim)
        if periodo:
            queryset = queryset.filter(periodo=periodo)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def ultimas_leituras(self, request):
        """Retorna as últimas leituras de todos os hidrômetros ativos"""
        hidrometros = Hidrometro.objects.filter(ativo=True)
        resultado = []
        
        for hidrometro in hidrometros:
            ultima_leitura = hidrometro.leituras.order_by('-data_leitura').first()
            if ultima_leitura:
                resultado.append({
                    'hidrometro': hidrometro.numero,
                    'lote': hidrometro.lote.numero,
                    'leitura': float(ultima_leitura.leitura),
                    'data_leitura': ultima_leitura.data_leitura,
                    'periodo': ultima_leitura.periodo
                })
        
        return Response(resultado)
    
    @action(detail=False, methods=['post'])
    def leitura_em_lote(self, request):
        """Permite criar múltiplas leituras de uma vez"""
        leituras_data = request.data.get('leituras', [])
        
        if not leituras_data:
            return Response(
                {'error': 'Nenhuma leitura fornecida'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        criadas = []
        erros = []
        
        for leitura_data in leituras_data:
            serializer = LeituraCreateSerializer(data=leitura_data)
            if serializer.is_valid():
                serializer.save()
                criadas.append(serializer.data)
            else:
                erros.append({
                    'dados': leitura_data,
                    'erros': serializer.errors
                })
        
        return Response({
            'criadas': len(criadas),
            'erros': len(erros),
            'leituras_criadas': criadas,
            'leituras_com_erro': erros
        }, status=status.HTTP_201_CREATED if criadas else status.HTTP_400_BAD_REQUEST)


# Views HTML para interface web
def dashboard(request):
    """Dashboard principal"""
    total_lotes = Lote.objects.filter(ativo=True).count()
    total_hidrometros = Hidrometro.objects.filter(ativo=True).count()
    
    hoje = timezone.now().date()
    leituras_hoje = Leitura.objects.filter(data_leitura__date=hoje).count()
    
    context = {
        'total_lotes': total_lotes,
        'total_hidrometros': total_hidrometros,
        'leituras_hoje': leituras_hoje,
    }
    
    return render(request, 'consumo/dashboard.html', context)


def listar_hidrometros(request):
    """Lista todos os hidrômetros com paginação"""
    from django.core.paginator import Paginator
    
    agora = timezone.localtime(timezone.now())
    hoje = agora.date()
    
    # Query base
    hidrometros_list = (
        Hidrometro.objects.filter(ativo=True)
        .select_related('lote')
        .annotate(
            leituras_hoje=Count('leituras', filter=Q(leituras__data_leitura__date=hoje)),
            ultima_leitura=Max('leituras__data_leitura'),
            # Cria um campo de ordenação: administração = 0, residencial = 1
            ordem_tipo=Case(
                When(lote__tipo='administracao', then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        )
    )
    
    # Filtro de busca
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Busca por número do hidrômetro OU número do lote
        hidrometros_list = hidrometros_list.filter(
            Q(numero__icontains=search_query) | 
            Q(lote__numero__icontains=search_query)
        )
    
    # Ordena: primeiro administração (ordem_tipo=0), depois residenciais (ordem_tipo=1), ambos por número crescente
    hidrometros_list = hidrometros_list.order_by('ordem_tipo', 'numero')
    
    # Paginação: 50 hidrômetros por página
    paginator = Paginator(hidrometros_list, 50)
    page_number = request.GET.get('page', 1)
    hidrometros = paginator.get_page(page_number)
    
    context = {
        'hidrometros': hidrometros,
        'search_query': search_query,
    }
    
    return render(request, 'consumo/listar_hidrometros.html', context)


def listar_leituras(request):
    """Lista todas as leituras com paginação"""
    from django.core.paginator import Paginator
    
    # Query base
    leituras_list = (
        Leitura.objects.all()
        .select_related('hidrometro__lote')
        .order_by('-data_leitura')
    )

    # Filtro de busca
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Busca por número do hidrômetro OU número do lote
        leituras_list = leituras_list.filter(
            Q(hidrometro__numero__icontains=search_query) | 
            Q(hidrometro__lote__numero__icontains=search_query)
        )

    # Paginação: 50 leituras por página
    paginator = Paginator(leituras_list, 50)
    page_number = request.GET.get('page', 1)
    leituras = paginator.get_page(page_number)
    total_leituras = paginator.count
    
    context = {
        'leituras': leituras,
        'total_leituras': total_leituras,
        'search_query': search_query,
    }
    
    return render(request, 'consumo/listar_leituras.html', context)


def registrar_leitura(request):
    """Formulário para registrar leituras"""
    hidrometros_queryset = Hidrometro.objects.filter(ativo=True).select_related('lote')

    def chave_natural(texto):
        import re
        partes = re.split(r'(\d+)', (texto or '').strip().lower())
        return [int(parte) if parte.isdigit() else parte for parte in partes]

    def chave_ordenacao_hidrometro(hidrometro):
        tipo_prioridade = 0 if hidrometro.lote.tipo == 'administracao' else 1
        if tipo_prioridade == 0:
            return (tipo_prioridade, chave_natural(hidrometro.numero), chave_natural(hidrometro.lote.numero))
        return (tipo_prioridade, chave_natural(hidrometro.lote.numero), chave_natural(hidrometro.numero))

    hidrometros = sorted(hidrometros_queryset, key=chave_ordenacao_hidrometro)
    
    context = {
        'hidrometros': hidrometros,
    }
    
    return render(request, 'consumo/registrar_leitura.html', context)


def detalhes_hidrometro(request, hidrometro_id):
    """Página com detalhes e histórico de leituras do hidrômetro com filtros e gráficos"""
    from datetime import timedelta
    from django.db.models import Sum
    from collections import defaultdict
    import calendar
    
    hidrometro = get_object_or_404(Hidrometro, id=hidrometro_id)
    
    # Obter filtros de período
    periodo = request.GET.get('periodo', '30dias')
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')
    
    hoje = timezone.now().date()
    data_inicio = None
    data_fim = hoje
    periodo_label = ''
    
    # Definir período baseado no filtro
    if periodo == '7dias':
        data_inicio = hoje - timedelta(days=7)
        periodo_label = 'Últimos 7 dias'
    elif periodo == '15dias':
        data_inicio = hoje - timedelta(days=15)
        periodo_label = 'Últimos 15 dias'
    elif periodo == '30dias':
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    elif periodo == 'mes_atual':
        data_inicio = hoje.replace(day=1)
        mes_nome = MESES_PT_BR[hoje.month]
        periodo_label = f'{mes_nome} de {hoje.year}'
    elif periodo == 'ano_atual':
        data_inicio = hoje.replace(month=1, day=1)
        periodo_label = f'Ano de {hoje.year}'
    elif periodo == 'personalizado' and data_inicio_str and data_fim_str:
        try:
            from datetime import datetime
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            periodo_label = f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
        except:
            data_inicio = hoje - timedelta(days=30)
            data_fim = hoje
            periodo_label = 'Últimos 30 dias'
            periodo = '30dias'
    else:
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    
    # Obter leituras filtradas
    leituras = hidrometro.leituras.filter(
        data_leitura__date__gte=data_inicio,
        data_leitura__date__lte=data_fim
    ).order_by('-data_leitura')
    
    # Obter todas as leituras para o histórico completo (limitado)
    leituras_historico = hidrometro.leituras.all().order_by('-data_leitura')[:50]
    
    # Calcular consumo total no período
    consumo_total_periodo = 0
    leituras_ordenadas = hidrometro.leituras.filter(
        data_leitura__date__gte=data_inicio,
        data_leitura__date__lte=data_fim
    ).order_by('data_leitura')
    
    for i, leitura_atual in enumerate(leituras_ordenadas):
        if i > 0:
            leitura_anterior = list(leituras_ordenadas)[i-1]
            diferenca = float(leitura_atual.leitura) - float(leitura_anterior.leitura)
            if diferenca > 0:
                consumo_litros = diferenca * 1000
                consumo_total_periodo += consumo_litros
    
    # Preparar dados para gráficos
    # Gráfico 1: Consumo por Dia
    consumo_por_dia = defaultdict(float)
    for i, leitura_atual in enumerate(leituras_ordenadas):
        if i > 0:
            leitura_anterior = list(leituras_ordenadas)[i-1]
            diferenca = float(leitura_atual.leitura) - float(leitura_anterior.leitura)
            if diferenca > 0:
                consumo_litros = diferenca * 1000
                dia_str = leitura_atual.data_leitura.strftime('%d/%m')
                consumo_por_dia[dia_str] += consumo_litros
    
    consumo_dia_lista = [
        {'dia': dia, 'consumo_litros': consumo}
        for dia, consumo in sorted(consumo_por_dia.items())
    ]
    
    # Gráfico 2: Consumo por Mês (sempre exibe todos os 12 meses)
    # Inicializar todos os meses com 0
    consumo_por_mes = {mes: 0.0 for mes in range(1, 13)}
    for i, leitura_atual in enumerate(leituras_ordenadas):
        if i > 0:
            leitura_anterior = list(leituras_ordenadas)[i-1]
            diferenca = float(leitura_atual.leitura) - float(leitura_anterior.leitura)
            if diferenca > 0:
                consumo_litros = diferenca * 1000
                mes_numero = leitura_atual.data_leitura.month
                consumo_por_mes[mes_numero] += consumo_litros
    
    # Sempre exibir todos os 12 meses em português
    consumo_mes_lista = []
    for mes in range(1, 13):
        mes_nome = MESES_PT_BR[mes]
        consumo_mes_lista.append({
            'mes': mes,
            'mes_nome': mes_nome,
            'consumo_litros': consumo_por_mes[mes]
        })
    
    # Dados dos gráficos (sem período do dia - removido do template)
    dados_graficos = {
        'consumo_dia': consumo_dia_lista,
        'consumo_mes': consumo_mes_lista,
        'consumo_total_periodo': consumo_total_periodo,
        'periodo_label': periodo_label,
        'periodo_selecionado': periodo,
    }
    
    # Serializar dados para JSON
    import json
    dados_graficos_json = json.dumps(dados_graficos, ensure_ascii=False)
    
    context = {
        'hidrometro': hidrometro,
        'leituras': leituras_historico,
        'dados_graficos': dados_graficos_json,
    }
    
    return render(request, 'consumo/detalhes_hidrometro.html', context)


def graficos_consumo(request):
    """Página com gráficos de consumo do condomínio com filtro de período."""

    agora = timezone.localtime(timezone.now())
    ultima_coleta = Leitura.objects.filter(
        hidrometro__ativo=True,
        hidrometro__lote__tipo='residencial'
    ).aggregate(ultima=Max('data_leitura'))['ultima']

    if ultima_coleta:
        hoje = timezone.localtime(ultima_coleta) if timezone.is_aware(ultima_coleta) else timezone.make_aware(ultima_coleta)
    else:
        hoje = agora

    ano_atual = hoje.year

    def _inicio_mes_menos(data_referencia, meses_anteriores):
        total_meses = data_referencia.year * 12 + (data_referencia.month - 1) - meses_anteriores
        ano = total_meses // 12
        mes = (total_meses % 12) + 1
        return data_referencia.replace(year=ano, month=mes, day=1, hour=0, minute=0, second=0, microsecond=0)

    # Obter o período selecionado (padrão: mês atual da última coleta)
    periodo_selecionado = request.GET.get('periodo', 'mes_atual')

    # Definir data de início baseada no período selecionado
    if periodo_selecionado == 'mes_atual':
        # Mês da última coleta (do dia 1 até a data da última coleta)
        data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_inicio_ano = data_inicio_dias
        periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    elif periodo_selecionado == '2meses':
        data_inicio_dias = _inicio_mes_menos(hoje, 1)
        data_inicio_ano = data_inicio_dias
        periodo_label = "Últimos 2 meses"
    elif periodo_selecionado == '3meses':
        data_inicio_dias = _inicio_mes_menos(hoje, 2)
        data_inicio_ano = data_inicio_dias
        periodo_label = "Últimos 3 meses"
    elif periodo_selecionado == 'ano_atual':
        # Ano atual (de 1º de janeiro até hoje)
        data_inicio_ano = timezone.datetime(ano_atual, 1, 1, 0, 0, 0, tzinfo=hoje.tzinfo)
        data_inicio_dias = data_inicio_ano
        periodo_label = f"Ano Atual ({ano_atual})"
    elif periodo_selecionado == 'personalizado':
        # Período personalizado (data_inicio e data_fim via GET)
        data_inicio_str = request.GET.get('data_inicio')
        data_fim_str = request.GET.get('data_fim')
        
        if data_inicio_str and data_fim_str:
            try:
                data_inicio_dias = timezone.datetime.strptime(data_inicio_str, '%Y-%m-%d')
                data_inicio_dias = timezone.make_aware(data_inicio_dias.replace(hour=0, minute=0, second=0, microsecond=0))
                data_fim_personalizada = timezone.datetime.strptime(data_fim_str, '%Y-%m-%d')
                data_fim_personalizada = timezone.make_aware(data_fim_personalizada.replace(hour=23, minute=59, second=59, microsecond=999999))
                
                # Limitar data_fim ao hoje se for futuro
                if data_fim_personalizada > hoje:
                    data_fim_personalizada = hoje
                
                data_inicio_ano = data_inicio_dias
                periodo_label = f"{data_inicio_dias.strftime('%d/%m/%Y')} até {data_fim_personalizada.strftime('%d/%m/%Y')}"
                hoje = data_fim_personalizada  # Usar data_fim personalizada
            except (ValueError, TypeError):
                # Se houver erro, usar padrão (mês atual da última coleta)
                data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                data_inicio_ano = data_inicio_dias
                periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
        else:
            # Sem datas fornecidas, usar padrão
            data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            data_inicio_ano = data_inicio_dias
            periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    else:
        # Padrão: mês atual da última coleta
        data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_inicio_ano = data_inicio_dias
        periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    
    data_fim = hoje

    dados_graficos = {
        'consumo_mes': [],
        'consumo_total_ano': 0.0,
        'top_lotes': [],
        'consumo_por_hidrometro': [],
        'periodo_label': periodo_label,
        'periodo_selecionado': periodo_selecionado,
        'ano_atual': ano_atual,
    }

    hidrometros_qs = Hidrometro.objects.filter(
        ativo=True,
        lote__tipo='residencial'
    ).select_related('lote')

    consumo_mensal = {mes: 0.0 for mes in range(1, 13)}
    consumo_total_ano = 0.0
    consumo_por_lote_ano = {}
    consumo_por_hidrometro = []

    for hidrometro in hidrometros_qs:
        # Buscar leituras do período filtrado para calcular o consumo total
        leituras_ano = hidrometro.leituras.filter(
            data_leitura__gte=data_inicio_dias,
            data_leitura__lte=data_fim
        ).order_by('data_leitura')

        consumo_hidrometro_litros = 0.0
        if not leituras_ano.exists():
            continue

        # Calcular consumo do ano (para total e top lotes)
        for i in range(1, len(leituras_ano)):
            leitura_atual = leituras_ano[i]
            leitura_anterior = leituras_ano[i - 1]

            consumo_m3 = float(leitura_atual.leitura - leitura_anterior.leitura)
            if consumo_m3 < 0:
                continue

            consumo_litros = consumo_m3 * 1000
            consumo_total_ano += consumo_litros
            consumo_hidrometro_litros += consumo_litros

            # Consumo por lote (ano)
            numero_lote = leitura_atual.hidrometro.lote.numero
            consumo_por_lote_ano.setdefault(numero_lote, 0.0)
            consumo_por_lote_ano[numero_lote] += consumo_litros

            # Consumo por mês do ano atual (sempre Jan-Dez no gráfico mensal)
            if leitura_atual.data_leitura.year == ano_atual:
                mes = leitura_atual.data_leitura.month
                consumo_mensal[mes] += consumo_litros

        if consumo_hidrometro_litros > 0:
            consumo_por_hidrometro.append({
                'hidrometro': hidrometro.numero,
                'lote': hidrometro.lote.numero,
                'consumo_litros': round(consumo_hidrometro_litros, 2),
            })

    nomes_meses = [
        'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
        'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'
    ]

    for mes in range(1, 13):
        dados_graficos['consumo_mes'].append({
            'mes': mes,
            'mes_nome': f"{nomes_meses[mes - 1]}/{str(ano_atual)[-2:]}",
            'consumo_litros': round(consumo_mensal[mes], 2)
        })

    dados_graficos['consumo_total_ano'] = round(consumo_total_ano, 2)

    top_lotes = sorted(consumo_por_lote_ano.items(), key=lambda x: x[1], reverse=True)[:10]
    dados_graficos['top_lotes'] = [
        {'lote': lote, 'consumo_litros': round(consumo, 2)} for lote, consumo in top_lotes
    ]

    def _ordenar_lote(item):
        numero = item['lote']
        # Lotes numéricos primeiro, ordenados por valor inteiro; depois lotes ADM/strings
        try:
            return (0, int(numero), numero)
        except ValueError:
            # Tentar extrair número após prefixo ADM-
            if numero.upper().startswith('ADM-'):
                try:
                    return (1, int(numero.split('-', 1)[1]), numero)
                except ValueError:
                    return (1, float('inf'), numero)
            return (1, float('inf'), numero)

    dados_graficos['consumo_por_hidrometro'] = sorted(
        consumo_por_hidrometro,
        key=lambda x: (_ordenar_lote(x), x['hidrometro'])
    )

    lotes_disponiveis = Lote.objects.filter(
        ativo=True,
        tipo='residencial'
    ).order_by('numero')

    context = {
        'dados_graficos': dados_graficos,
        'hidrometros': hidrometros_qs,
        'lotes': lotes_disponiveis,
    }

    return render(request, 'consumo/graficos_consumo.html', context)


def graficos_lote(request, lote_id):
    """Página com gráficos de consumo específicos de um lote com filtros de período"""
    from collections import defaultdict
    import calendar
    
    lote = get_object_or_404(Lote, id=lote_id)
    
    # Obter filtros de período
    periodo = request.GET.get('periodo', '30dias')
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')
    
    hoje = timezone.now().date()
    data_inicio = None
    data_fim = hoje
    periodo_label = ''
    
    # Definir período baseado no filtro
    if periodo == '7dias':
        data_inicio = hoje - timedelta(days=7)
        periodo_label = 'Últimos 7 dias'
    elif periodo == '15dias':
        data_inicio = hoje - timedelta(days=15)
        periodo_label = 'Últimos 15 dias'
    elif periodo == '30dias':
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    elif periodo == 'mes_atual':
        data_inicio = hoje.replace(day=1)
        mes_nome = MESES_PT_BR[hoje.month]
        periodo_label = f'{mes_nome} de {hoje.year}'
    elif periodo == 'ano_atual':
        data_inicio = hoje.replace(month=1, day=1)
        periodo_label = f'Ano de {hoje.year}'
    elif periodo == 'personalizado' and data_inicio_str and data_fim_str:
        try:
            from datetime import datetime
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            periodo_label = f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
        except:
            data_inicio = hoje - timedelta(days=30)
            data_fim = hoje
            periodo_label = 'Últimos 30 dias'
            periodo = '30dias'
    else:
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    
    # Obter todos os hidrômetros do lote
    hidrometros = lote.hidrometros.filter(ativo=True)
    
    if not hidrometros.exists():
        dados_graficos = {
            'lote': lote.numero,
            'tipo': lote.get_tipo_display(),
            'consumo_por_dia': [],
            'consumo_mes': [],
            'consumo_periodo': [],
            'consumo_total_periodo': 0,
            'periodo_label': periodo_label,
            'periodo_selecionado': periodo,
        }
        
        import json
        dados_graficos_json = json.dumps(dados_graficos, ensure_ascii=False)
        
        context = {
            'lote': lote,
            'dados_graficos': dados_graficos_json,
            'sem_dados': True,
        }
        return render(request, 'consumo/graficos_lote.html', context)
    
    # Calcular consumo total no período para todos os hidrômetros do lote
    consumo_total_periodo = 0
    
    for hidrometro in hidrometros:
        leituras_ordenadas = hidrometro.leituras.filter(
            data_leitura__date__gte=data_inicio,
            data_leitura__date__lte=data_fim
        ).order_by('data_leitura')
        
        for i, leitura_atual in enumerate(leituras_ordenadas):
            if i > 0:
                leitura_anterior = list(leituras_ordenadas)[i-1]
                diferenca = float(leitura_atual.leitura) - float(leitura_anterior.leitura)
                if diferenca > 0:
                    consumo_litros = diferenca * 1000
                    consumo_total_periodo += consumo_litros
    
    # Preparar dados para gráficos
    # Gráfico 1: Consumo por Dia
    consumo_por_dia = defaultdict(float)
    
    for hidrometro in hidrometros:
        leituras_ordenadas = hidrometro.leituras.filter(
            data_leitura__date__gte=data_inicio,
            data_leitura__date__lte=data_fim
        ).order_by('data_leitura')
        
        for i, leitura_atual in enumerate(leituras_ordenadas):
            if i > 0:
                leitura_anterior = list(leituras_ordenadas)[i-1]
                diferenca = float(leitura_atual.leitura) - float(leitura_anterior.leitura)
                if diferenca > 0:
                    consumo_litros = diferenca * 1000
                    dia_str = leitura_atual.data_leitura.strftime('%d/%m')
                    consumo_por_dia[dia_str] += consumo_litros
    
    consumo_dia_lista = [
        {'dia': dia, 'consumo_litros': consumo}
        for dia, consumo in sorted(consumo_por_dia.items())
    ]
    
    # Gráfico 2: Consumo por Mês (sempre exibe todos os 12 meses)
    # Inicializar todos os meses com 0
    consumo_por_mes = {mes: 0.0 for mes in range(1, 13)}
    
    for hidrometro in hidrometros:
        leituras_ordenadas = hidrometro.leituras.filter(
            data_leitura__date__gte=data_inicio,
            data_leitura__date__lte=data_fim
        ).order_by('data_leitura')
        
        for i, leitura_atual in enumerate(leituras_ordenadas):
            if i > 0:
                leitura_anterior = list(leituras_ordenadas)[i-1]
                diferenca = float(leitura_atual.leitura) - float(leitura_anterior.leitura)
                if diferenca > 0:
                    consumo_litros = diferenca * 1000
                    mes_numero = leitura_atual.data_leitura.month
                    consumo_por_mes[mes_numero] += consumo_litros
    
    # Sempre exibir todos os 12 meses em português
    consumo_mes_lista = []
    for mes in range(1, 13):
        mes_nome = MESES_PT_BR[mes]
        consumo_mes_lista.append({
            'mes': mes,
            'mes_nome': mes_nome,
            'consumo_litros': consumo_por_mes[mes]
        })
    
    # Dados dos gráficos (sem período do dia - removido do template)
    dados_graficos = {
        'lote': lote.numero,
        'tipo': lote.get_tipo_display(),
        'consumo_por_dia': consumo_dia_lista,
        'consumo_mes': consumo_mes_lista,
        'consumo_total_periodo': consumo_total_periodo,
        'periodo_label': periodo_label,
        'periodo_selecionado': periodo,
    }
    
    # Serializar dados para JSON
    import json
    dados_graficos_json = json.dumps(dados_graficos, ensure_ascii=False)
    
    context = {
        'lote': lote,
        'dados_graficos': dados_graficos_json,
        'hidrometros': hidrometros,
    }
    
    return render(request, 'consumo/graficos_lote.html', context)


def exportar_graficos_consumo_pdf(request):
    """Exporta os gráficos de consumo do condomínio em PDF"""
    import os
    os.environ.setdefault('MPLCONFIGDIR', '/tmp/matplotlib')
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    
    from django.template.loader import render_to_string
    
    # Obter dados dos gráficos (mesma lógica da view graficos_consumo)
    agora = timezone.localtime(timezone.now())
    ultima_coleta = Leitura.objects.filter(
        hidrometro__ativo=True,
        hidrometro__lote__tipo='residencial'
    ).aggregate(ultima=Max('data_leitura'))['ultima']

    if ultima_coleta:
        hoje = timezone.localtime(ultima_coleta) if timezone.is_aware(ultima_coleta) else timezone.make_aware(ultima_coleta)
    else:
        hoje = agora

    ano_atual = hoje.year

    def _inicio_mes_menos(data_referencia, meses_anteriores):
        total_meses = data_referencia.year * 12 + (data_referencia.month - 1) - meses_anteriores
        ano = total_meses // 12
        mes = (total_meses % 12) + 1
        return data_referencia.replace(year=ano, month=mes, day=1, hour=0, minute=0, second=0, microsecond=0)

    # Obter o período selecionado (padrão: mês atual da última coleta)
    periodo_selecionado = request.GET.get('periodo', 'mes_atual')

    # Definir data de início baseada no período selecionado
    if periodo_selecionado == 'mes_atual':
        # Mês da última coleta (do dia 1 até a data da última coleta)
        data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_inicio_ano = data_inicio_dias
        periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    elif periodo_selecionado == '2meses':
        data_inicio_dias = _inicio_mes_menos(hoje, 1)
        data_inicio_ano = data_inicio_dias
        periodo_label = "Últimos 2 meses"
    elif periodo_selecionado == '3meses':
        data_inicio_dias = _inicio_mes_menos(hoje, 2)
        data_inicio_ano = data_inicio_dias
        periodo_label = "Últimos 3 meses"
    elif periodo_selecionado == 'ano_atual':
        # Ano atual (de 1º de janeiro até hoje)
        data_inicio_ano = timezone.datetime(ano_atual, 1, 1, 0, 0, 0, tzinfo=hoje.tzinfo)
        data_inicio_dias = data_inicio_ano
        periodo_label = f"Ano Atual ({ano_atual})"
    elif periodo_selecionado == 'personalizado':
        # Período personalizado (data_inicio e data_fim via GET)
        data_inicio_str = request.GET.get('data_inicio')
        data_fim_str = request.GET.get('data_fim')
        
        if data_inicio_str and data_fim_str:
            try:
                data_inicio_dias = timezone.datetime.strptime(data_inicio_str, '%Y-%m-%d')
                data_inicio_dias = timezone.make_aware(data_inicio_dias.replace(hour=0, minute=0, second=0, microsecond=0))
                data_fim_personalizada = timezone.datetime.strptime(data_fim_str, '%Y-%m-%d')
                data_fim_personalizada = timezone.make_aware(data_fim_personalizada.replace(hour=23, minute=59, second=59, microsecond=999999))
                
                # Limitar data_fim ao hoje se for futuro
                if data_fim_personalizada > hoje:
                    data_fim_personalizada = hoje
                
                data_inicio_ano = data_inicio_dias
                periodo_label = f"{data_inicio_dias.strftime('%d/%m/%Y')} até {data_fim_personalizada.strftime('%d/%m/%Y')}"
                hoje = data_fim_personalizada  # Usar data_fim personalizada
            except (ValueError, TypeError):
                # Se houver erro, usar padrão (mês atual da última coleta)
                data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                data_inicio_ano = data_inicio_dias
                periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
        else:
            # Sem datas fornecidas, usar padrão
            data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            data_inicio_ano = data_inicio_dias
            periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    else:
        # Padrão: mês atual da última coleta
        data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_inicio_ano = data_inicio_dias
        periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    
    data_fim = hoje
    
    # Buscar todos os hidrômetros ativos
    hidrometros = Hidrometro.objects.filter(
        ativo=True,
        lote__tipo='residencial'
    ).select_related('lote')
    
    # Consumo por hidrômetro (individual) no período
    consumo_por_hidrometro = []
    consumo_total_periodo = 0.0
    
    for hidrometro in hidrometros:
        leituras = hidrometro.leituras.filter(
            data_leitura__gte=data_inicio_dias,
            data_leitura__lte=data_fim
        ).order_by('data_leitura')
        
        consumo_hidrometro_litros = 0.0
        if leituras.exists():
            for i in range(1, len(leituras)):
                leitura_atual = leituras[i]
                leitura_anterior = leituras[i - 1]
                
                consumo_m3 = float(leitura_atual.leitura - leitura_anterior.leitura)
                if consumo_m3 < 0:
                    continue
                    
                consumo_litros = consumo_m3 * 1000
                consumo_hidrometro_litros += consumo_litros
                consumo_total_periodo += consumo_litros
                
        if consumo_hidrometro_litros > 0:
            consumo_por_hidrometro.append({
                'hidrometro': hidrometro.numero,
                'lote': hidrometro.lote.numero,
                'consumo_litros': round(consumo_hidrometro_litros, 2),
            })
    
    # Top 10 lotes por consumo (baseado no período filtrado)
    lotes_consumo = []
    for lote in Lote.objects.filter(ativo=True, tipo='residencial'):
        consumo_lote = 0.0
        hidrometros_lote = lote.hidrometros.filter(ativo=True)
        
        for hidrometro in hidrometros_lote:
            leituras_periodo = hidrometro.leituras.filter(
                data_leitura__gte=data_inicio_dias,
                data_leitura__lte=data_fim
            ).order_by('data_leitura')
            
            if leituras_periodo.count() >= 2:
                primeira = leituras_periodo.first()
                ultima = leituras_periodo.last()
                consumo_m3 = float(ultima.leitura - primeira.leitura)
                consumo_litros = consumo_m3 * 1000
                consumo_lote += consumo_litros
        
        if consumo_lote > 0:
            lotes_consumo.append({
                'lote': lote,
                'consumo': consumo_lote
            })
    
    lotes_consumo.sort(key=lambda x: x['consumo'], reverse=True)
    top_lotes = lotes_consumo[:10]

    # Ordenar hidrômetros por lote (numéricos primeiro, depois ADM)
    def _ordenar_lote(item):
        numero = item['lote']
        try:
            return (0, int(numero), numero)
        except ValueError:
            if numero.upper().startswith('ADM-'):
                try:
                    return (1, int(numero.split('-', 1)[1]), numero)
                except ValueError:
                    return (1, float('inf'), numero)
            return (1, float('inf'), numero)

    consumo_por_hidrometro = sorted(
        consumo_por_hidrometro,
        key=lambda x: (_ordenar_lote(x), x['hidrometro'])
    )
    
    # Criar PDF
    img_buffer = None
    img_buffer_top = None
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                          rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo do título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#7f8c8d'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Título
    elements.append(Paragraph(f"Relatório de Consumo de Água - {periodo_label}", title_style))
    elements.append(Paragraph(f"Gerado em: {agora.strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumo Geral
    elements.append(Paragraph("📊 Resumo Geral", heading_style))
    
    resumo_data = [
        ['Indicador', 'Valor'],
        ['Período', periodo_label],
        ['Consumo Total', f'{consumo_total_periodo:,.0f} L'],
        ['Hidrômetros Ativos', str(hidrometros.count())],
        ['Lotes Ativos', str(Lote.objects.filter(ativo=True, tipo='residencial').count())],
    ]
    
    resumo_table = Table(resumo_data, colWidths=[3*inch, 2*inch])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Top 10 Lotes
    elements.append(Paragraph("🏆 Top 10 Lotes com Maior Consumo", heading_style))
    
    top_data = [['Posição', 'Lote', 'Tipo', 'Consumo (L)']]
    for idx, item in enumerate(top_lotes, 1):
        lote = item['lote']
        consumo = item['consumo']
        top_data.append([
            str(idx),
            lote.numero,
            lote.get_tipo_display(),
            f'{consumo:,.2f}'
        ])
    
    top_table = Table(top_data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 2*inch])
    top_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(top_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Gráfico Top 10 Lotes
    if top_lotes:
        plt.figure(figsize=(10, 5))
        lotes_labels = [item['lote'].numero for item in top_lotes]
        lotes_valores = [item['consumo'] for item in top_lotes]
        plt.barh(lotes_labels[::-1], lotes_valores[::-1], color='#e74c3c', alpha=0.7)
        plt.title(f'Top 10 Lotes - Consumo ({periodo_label})', fontsize=14, fontweight='bold')
        plt.xlabel('Consumo (L)', fontsize=11)
        plt.ylabel('Lote', fontsize=11)
        plt.grid(axis='x', alpha=0.3)
        plt.tight_layout()
        
        # Salvar gráfico em buffer
        img_buffer_top = io.BytesIO()
        plt.savefig(img_buffer_top, format='png', dpi=150, bbox_inches='tight')
        img_buffer_top.seek(0)
        plt.close()
        
        # Adicionar imagem ao PDF
        img_top = Image(img_buffer_top, width=7*inch, height=3.5*inch)
        elements.append(img_top)
    
    elements.append(Spacer(1, 0.3*inch))
    elements.append(PageBreak())
    
    # Tabela: Consumo por Hidrômetro (período)
    elements.append(Paragraph("📈 Consumo por Hidrômetro (período)", heading_style))

    hidrometro_data = [['Hidrômetro', 'Lote', 'Consumo (L)']]
    for item in consumo_por_hidrometro:
        hidrometro_data.append([
            item['hidrometro'],
            item['lote'],
            f"{item['consumo_litros']:,.0f}"
        ])

    hidrometro_table = Table(hidrometro_data, colWidths=[2*inch, 1.5*inch, 2*inch])
    hidrometro_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2980b9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))

    elements.append(hidrometro_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Construir PDF
    doc.build(
        elements,
        onFirstPage=_desenhar_marca_dagua_logo,
        onLaterPages=_desenhar_marca_dagua_logo
    )
    
    # Preparar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_consumo_condominio_{agora.strftime("%Y%m%d")}.pdf"'
    buffer.close()
    if img_buffer is not None:
        img_buffer.close()
    if img_buffer_top is not None:
        img_buffer_top.close()
    
    return response


def baixar_relatorios_lotes_periodo_zip(request):
    """Baixa todos os relatórios individuais de lotes (com fotos embutidas no PDF) em um único ZIP."""
    agora = timezone.localtime(timezone.now())
    ultima_coleta = Leitura.objects.filter(
        hidrometro__ativo=True,
        hidrometro__lote__tipo='residencial'
    ).aggregate(ultima=Max('data_leitura'))['ultima']

    if ultima_coleta:
        hoje_ref = timezone.localtime(ultima_coleta) if timezone.is_aware(ultima_coleta) else timezone.make_aware(ultima_coleta)
    else:
        hoje_ref = agora

    periodo_selecionado = request.GET.get('periodo', 'mes_atual')

    def _inicio_mes_menos(data_referencia, meses_anteriores):
        total_meses = data_referencia.year * 12 + (data_referencia.month - 1) - meses_anteriores
        ano = total_meses // 12
        mes = (total_meses % 12) + 1
        return data_referencia.replace(year=ano, month=mes, day=1, hour=0, minute=0, second=0, microsecond=0)

    if periodo_selecionado == 'mes_atual':
        data_inicio_dt = hoje_ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_fim_dt = hoje_ref
    elif periodo_selecionado == '2meses':
        data_inicio_dt = _inicio_mes_menos(hoje_ref, 1)
        data_fim_dt = hoje_ref
    elif periodo_selecionado == '3meses':
        data_inicio_dt = _inicio_mes_menos(hoje_ref, 2)
        data_fim_dt = hoje_ref
    elif periodo_selecionado == 'ano_atual':
        data_inicio_dt = timezone.datetime(hoje_ref.year, 1, 1, 0, 0, 0, tzinfo=hoje_ref.tzinfo)
        data_fim_dt = hoje_ref
    elif periodo_selecionado == 'personalizado':
        data_inicio_str = request.GET.get('data_inicio')
        data_fim_str = request.GET.get('data_fim')
        if data_inicio_str and data_fim_str:
            try:
                data_inicio_dt = timezone.make_aware(
                    timezone.datetime.strptime(data_inicio_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
                )
                data_fim_dt = timezone.make_aware(
                    timezone.datetime.strptime(data_fim_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
                )
                if data_fim_dt > hoje_ref:
                    data_fim_dt = hoje_ref
            except (ValueError, TypeError):
                data_inicio_dt = hoje_ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                data_fim_dt = hoje_ref
        else:
            data_inicio_dt = hoje_ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            data_fim_dt = hoje_ref
    else:
        data_inicio_dt = hoje_ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_fim_dt = hoje_ref

    data_inicio = data_inicio_dt.date()
    data_fim = data_fim_dt.date()
    intervalo_token = f"{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}"

    pasta_relatorios = os.path.join(settings.BASE_DIR, f'relatorios_lotes_{intervalo_token}')
    pasta_pacote_zip = f'pacote_relatorios_lotes_{intervalo_token}'
    subpasta_relatorios = f'relatorios_lotes_{intervalo_token}'

    buffer = io.BytesIO()
    total_arquivos = 0
    nomes_zip_adicionados = set()

    with zipfile.ZipFile(buffer, 'w', compression=zipfile.ZIP_DEFLATED) as arquivo_zip:
        if os.path.isdir(pasta_relatorios):
            for raiz, _, arquivos in os.walk(pasta_relatorios):
                for nome_arquivo in arquivos:
                    if not nome_arquivo.lower().endswith('.pdf'):
                        continue
                    caminho_arquivo = os.path.join(raiz, nome_arquivo)
                    relativo_relatorios = os.path.relpath(caminho_arquivo, pasta_relatorios)
                    caminho_zip = os.path.join(
                        pasta_pacote_zip,
                        subpasta_relatorios,
                        relativo_relatorios,
                    )
                    if caminho_zip in nomes_zip_adicionados:
                        continue
                    arquivo_zip.write(caminho_arquivo, caminho_zip)
                    nomes_zip_adicionados.add(caminho_zip)
                    total_arquivos += 1

        padrao_raiz = os.path.join(
            settings.BASE_DIR,
            f'relatorio_lote_*_{data_inicio.strftime("%Y%m%d")}_{data_fim.strftime("%Y%m%d")}.pdf'
        )
        for caminho_pdf in glob.glob(padrao_raiz):
            nome_pdf = os.path.basename(caminho_pdf)
            caminho_zip = os.path.join(pasta_pacote_zip, subpasta_relatorios, nome_pdf)
            if caminho_zip in nomes_zip_adicionados:
                continue
            arquivo_zip.write(caminho_pdf, caminho_zip)
            nomes_zip_adicionados.add(caminho_zip)
            total_arquivos += 1

    if total_arquivos == 0:
        comando_geracao = (
            f"python manage.py gerar_relatorios_lotes_periodo "
            f"--data-inicio {data_inicio.strftime('%Y-%m-%d')} "
            f"--data-fim {data_fim.strftime('%Y-%m-%d')}"
        )
        return HttpResponse(
            (
                f'Nenhum relatório PDF foi encontrado para o período de '
                f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}. '
                f'Gere os relatórios com: {comando_geracao} e tente novamente.'
            ),
            status=404,
        )

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = (
        f'attachment; filename="pacote_relatorios_lotes_{intervalo_token}.zip"'
    )
    return response


def exportar_graficos_consumo_excel(request):
    """Exporta os gráficos de consumo do condomínio em Excel com gráficos"""
    import os
    os.environ.setdefault('MPLCONFIGDIR', '/tmp/matplotlib')
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    
    # Obter dados dos gráficos (mesma lógica da view graficos_consumo)
    agora = timezone.localtime(timezone.now())
    ultima_coleta = Leitura.objects.filter(
        hidrometro__ativo=True,
        hidrometro__lote__tipo='residencial'
    ).aggregate(ultima=Max('data_leitura'))['ultima']

    if ultima_coleta:
        hoje = timezone.localtime(ultima_coleta) if timezone.is_aware(ultima_coleta) else timezone.make_aware(ultima_coleta)
    else:
        hoje = agora

    ano_atual = hoje.year

    def _inicio_mes_menos(data_referencia, meses_anteriores):
        total_meses = data_referencia.year * 12 + (data_referencia.month - 1) - meses_anteriores
        ano = total_meses // 12
        mes = (total_meses % 12) + 1
        return data_referencia.replace(year=ano, month=mes, day=1, hour=0, minute=0, second=0, microsecond=0)

    # Obter o período selecionado (padrão: mês atual da última coleta)
    periodo_selecionado = request.GET.get('periodo', 'mes_atual')

    # Definir data de início baseada no período selecionado
    if periodo_selecionado == 'mes_atual':
        # Mês da última coleta (do dia 1 até a data da última coleta)
        data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_inicio_ano = data_inicio_dias
        periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    elif periodo_selecionado == '2meses':
        data_inicio_dias = _inicio_mes_menos(hoje, 1)
        data_inicio_ano = data_inicio_dias
        periodo_label = "Últimos 2 meses"
    elif periodo_selecionado == '3meses':
        data_inicio_dias = _inicio_mes_menos(hoje, 2)
        data_inicio_ano = data_inicio_dias
        periodo_label = "Últimos 3 meses"
    elif periodo_selecionado == 'ano_atual':
        # Ano atual (de 1º de janeiro até hoje)
        data_inicio_ano = timezone.datetime(ano_atual, 1, 1, 0, 0, 0, tzinfo=hoje.tzinfo)
        data_inicio_dias = data_inicio_ano
        periodo_label = f"Ano Atual ({ano_atual})"
    elif periodo_selecionado == 'personalizado':
        # Período personalizado (data_inicio e data_fim via GET)
        data_inicio_str = request.GET.get('data_inicio')
        data_fim_str = request.GET.get('data_fim')
        
        if data_inicio_str and data_fim_str:
            try:
                data_inicio_dias = timezone.datetime.strptime(data_inicio_str, '%Y-%m-%d')
                data_inicio_dias = timezone.make_aware(data_inicio_dias.replace(hour=0, minute=0, second=0, microsecond=0))
                data_fim_personalizada = timezone.datetime.strptime(data_fim_str, '%Y-%m-%d')
                data_fim_personalizada = timezone.make_aware(data_fim_personalizada.replace(hour=23, minute=59, second=59, microsecond=999999))
                
                # Limitar data_fim ao hoje se for futuro
                if data_fim_personalizada > hoje:
                    data_fim_personalizada = hoje
                
                data_inicio_ano = data_inicio_dias
                periodo_label = f"{data_inicio_dias.strftime('%d/%m/%Y')} até {data_fim_personalizada.strftime('%d/%m/%Y')}"
                hoje = data_fim_personalizada  # Usar data_fim personalizada
            except (ValueError, TypeError):
                # Se houver erro, usar padrão (mês atual da última coleta)
                data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                data_inicio_ano = data_inicio_dias
                periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
        else:
            # Sem datas fornecidas, usar padrão
            data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            data_inicio_ano = data_inicio_dias
            periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    else:
        # Padrão: mês atual da última coleta
        data_inicio_dias = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        data_inicio_ano = data_inicio_dias
        periodo_label = f"Mês Atual ({formatar_mes_ano_ptbr(hoje)}) - Última coleta"
    
    data_fim = hoje
    
    # Buscar todos os hidrômetros ativos
    hidrometros = Hidrometro.objects.filter(
        ativo=True,
        lote__tipo='residencial'
    ).select_related('lote')
    
    # Consumo por hidrômetro (individual) no período
    consumo_por_hidrometro = []
    consumo_total_periodo = 0.0
    
    for hidrometro in hidrometros:
        leituras = hidrometro.leituras.filter(
            data_leitura__gte=data_inicio_dias,
            data_leitura__lte=data_fim
        ).order_by('data_leitura')
        
        consumo_hidrometro_litros = 0.0
        if leituras.exists():
            for i in range(1, len(leituras)):
                leitura_atual = leituras[i]
                leitura_anterior = leituras[i - 1]
                
                consumo_m3 = float(leitura_atual.leitura - leitura_anterior.leitura)
                if consumo_m3 < 0:
                    continue
                    
                consumo_litros = consumo_m3 * 1000
                consumo_hidrometro_litros += consumo_litros
                consumo_total_periodo += consumo_litros
                
        consumo_por_hidrometro.append({
            'hidrometro': hidrometro.numero,
            'lote': hidrometro.lote.numero,
            'consumo_litros': round(consumo_hidrometro_litros, 2),
        })
    
    # Top 10 lotes por consumo (baseado no período filtrado)
    lotes_consumo = []
    for lote in Lote.objects.filter(ativo=True, tipo='residencial'):
        consumo_lote = 0.0
        hidrometros_lote = lote.hidrometros.filter(ativo=True)
        
        for hidrometro in hidrometros_lote:
            leituras_periodo = hidrometro.leituras.filter(
                data_leitura__gte=data_inicio_dias,
                data_leitura__lte=data_fim
            ).order_by('data_leitura')
            
            if leituras_periodo.count() >= 2:
                primeira = leituras_periodo.first()
                ultima = leituras_periodo.last()
                consumo_m3 = float(ultima.leitura - primeira.leitura)
                consumo_litros = consumo_m3 * 1000
                consumo_lote += consumo_litros
        
        if consumo_lote > 0:
            lotes_consumo.append({
                'lote': lote,
                'consumo': consumo_lote
            })
    
    lotes_consumo.sort(key=lambda x: x['consumo'], reverse=True)
    top_lotes = lotes_consumo[:10]
    
    # Ordenar hidrômetros por lote (numéricos primeiro, depois ADM)
    def _ordenar_lote(item):
        numero = item['lote']
        try:
            return (0, int(numero), numero)
        except ValueError:
            if numero.upper().startswith('ADM-'):
                try:
                    return (1, int(numero.split('-', 1)[1]), numero)
                except ValueError:
                    return (1, float('inf'), numero)
            return (1, float('inf'), numero)

    consumo_por_hidrometro = sorted(
        consumo_por_hidrometro,
        key=lambda x: (_ordenar_lote(x), x['hidrometro'])
    )
    
    # Criar Excel
    wb = Workbook()
    
    # Aba: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Título
    ws_resumo['A1'] = f'Relatório de Consumo de Água - {periodo_label}'
    ws_resumo['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_resumo['A1'].fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    ws_resumo['A1'].alignment = Alignment(horizontal='center')
    ws_resumo.merge_cells('A1:C1')
    
    ws_resumo['A2'] = f'Gerado em: {agora.strftime("%d/%m/%Y %H:%M")}'
    ws_resumo['A2'].alignment = Alignment(horizontal='center')
    ws_resumo.merge_cells('A2:C2')
    
    # Dados resumo
    ws_resumo['A4'] = 'Indicador'
    ws_resumo['B4'] = 'Valor'
    ws_resumo['A4'].font = Font(bold=True)
    ws_resumo['B4'].font = Font(bold=True)
    
    resumo_dados = [
        ['Período', periodo_label],
        ['Consumo Total', f'{consumo_total_periodo:,.0f} L'],
        ['Hidrômetros Ativos', hidrometros.count()],
        ['Lotes Ativos', Lote.objects.filter(ativo=True, tipo='residencial').count()],
    ]
    
    for idx, (indicador, valor) in enumerate(resumo_dados, start=5):
        ws_resumo[f'A{idx}'] = indicador
        ws_resumo[f'B{idx}'] = valor
    
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 20
    
    # Aba: Top 10 Lotes
    ws_top = wb.create_sheet("Top 10 Lotes")
    
    ws_top['A1'] = 'Posição'
    ws_top['B1'] = 'Lote'
    ws_top['C1'] = 'Tipo'
    ws_top['D1'] = 'Consumo (L)'
    
    for col in ['A1', 'B1', 'C1', 'D1']:
        ws_top[col].font = Font(bold=True)
    
    for idx, item in enumerate(top_lotes, 1):
        lote = item['lote']
        consumo = item['consumo']
        ws_top[f'A{idx + 1}'] = idx
        ws_top[f'B{idx + 1}'] = lote.numero
        ws_top[f'C{idx + 1}'] = lote.get_tipo_display()
        ws_top[f'D{idx + 1}'] = round(consumo, 2)
    
    # Gerar gráfico com matplotlib (igual ao PDF)
    if top_lotes:
        plt.figure(figsize=(12, 6))
        lotes_labels = [item['lote'].numero for item in top_lotes]
        lotes_valores = [item['consumo'] for item in top_lotes]
        plt.barh(lotes_labels[::-1], lotes_valores[::-1], color='#e74c3c', alpha=0.7)
        plt.title(f'Top 10 Lotes - Consumo ({periodo_label})', fontsize=14, fontweight='bold')
        plt.xlabel('Consumo (L)', fontsize=11)
        plt.ylabel('Lote', fontsize=11)
        plt.grid(axis='x', alpha=0.3)
        plt.tight_layout()
        
        # Salvar gráfico em buffer
        img_buffer_top = io.BytesIO()
        plt.savefig(img_buffer_top, format='png', dpi=100, bbox_inches='tight')
        img_buffer_top.seek(0)
        plt.close()
        
        # Adicionar imagem ao Excel
        from openpyxl.drawing.image import Image as XLImage
        img_top_chart = XLImage(img_buffer_top)
        img_top_chart.width = 600
        img_top_chart.height = 300
        ws_top.add_image(img_top_chart, "F2")
    
    for col in ['A', 'B', 'C', 'D']:
        ws_top.column_dimensions[col].width = 15

    # Aba: Consumo por Hidrômetro
    ws_hid = wb.create_sheet("Consumo por Hidrômetro")
    ws_hid['A1'] = 'Hidrômetro'
    ws_hid['B1'] = 'Lote'
    ws_hid['C1'] = 'Consumo (L)'
    for col in ['A1', 'B1', 'C1']:
        ws_hid[col].font = Font(bold=True)

    for idx, item in enumerate(consumo_por_hidrometro, start=2):
        ws_hid[f'A{idx}'] = item['hidrometro']
        ws_hid[f'B{idx}'] = item['lote']
        ws_hid[f'C{idx}'] = item['consumo_litros']

    for col in ['A', 'B', 'C']:
        ws_hid.column_dimensions[col].width = 18

    # Gráfico de barras por hidrômetro
    if consumo_por_hidrometro:
        plt.figure(figsize=(14, 6))
        labels_h = [f"{item['hidrometro']} (Lote {item['lote']})" for item in consumo_por_hidrometro]
        valores_h = [item['consumo_litros'] for item in consumo_por_hidrometro]
        plt.bar(range(len(labels_h)), valores_h, color='#eab308', alpha=0.85)
        plt.title(f'Consumo por Hidrômetro ({periodo_label})', fontsize=14, fontweight='bold')
        plt.xlabel('Hidrômetro', fontsize=11)
        plt.ylabel('Consumo (L)', fontsize=11)
        plt.xticks(range(len(labels_h)), labels_h, rotation=60, ha='right', fontsize=8)
        plt.grid(axis='y', alpha=0.3)
        plt.tight_layout()

        img_buffer_h = io.BytesIO()
        plt.savefig(img_buffer_h, format='png', dpi=100, bbox_inches='tight')
        img_buffer_h.seek(0)
        plt.close()

        from openpyxl.drawing.image import Image as XLImage
        img_h_chart = XLImage(img_buffer_h)
        img_h_chart.width = 700
        img_h_chart.height = 320
        ws_hid.add_image(img_h_chart, "E2")
    
    # Salvar e retornar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_consumo_condominio_{agora.strftime("%Y%m%d")}.xlsx"'
    
    return response


def exportar_graficos_lote_pdf(request, lote_id):
    """Exporta os gráficos de consumo de um lote específico em PDF"""
    import os
    os.environ.setdefault('MPLCONFIGDIR', '/tmp/matplotlib')
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    
    lote = get_object_or_404(Lote, id=lote_id)
    
    # Obter dados do lote (mesma lógica da view graficos_lote)
    agora = timezone.localtime(timezone.now())
    hoje = agora.date()
    periodo = request.GET.get('periodo', '30dias')
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')
    data_fim = hoje
    periodo_label = ''

    if periodo == '7dias':
        data_inicio = hoje - timedelta(days=7)
        periodo_label = 'Últimos 7 dias'
    elif periodo == '15dias':
        data_inicio = hoje - timedelta(days=15)
        periodo_label = 'Últimos 15 dias'
    elif periodo == '30dias':
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    elif periodo == 'mes_atual':
        data_inicio = hoje.replace(day=1)
        mes_nome = MESES_PT_BR[hoje.month]
        periodo_label = f'{mes_nome} de {hoje.year}'
    elif periodo == 'ano_atual':
        data_inicio = hoje.replace(month=1, day=1)
        periodo_label = f'Ano de {hoje.year}'
    elif periodo == 'personalizado' and data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            if data_fim > hoje:
                data_fim = hoje
            periodo_label = f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
        except (ValueError, TypeError):
            data_inicio = hoje - timedelta(days=30)
            data_fim = hoje
            periodo_label = 'Últimos 30 dias'
            periodo = '30dias'
    else:
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    
    hidrometros = lote.hidrometros.filter(ativo=True)
    
    if not hidrometros.exists():
        return HttpResponse("Nenhum hidrômetro ativo encontrado para este lote.", status=404)
    
    # Calcular consumo no periodo
    nomes_meses = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]
    consumo_total_periodo = 0.0
    consumo_por_dia = {}
    consumo_por_mes = {}

    for hidrometro in hidrometros:
        leituras = hidrometro.leituras.filter(
            data_leitura__date__gte=data_inicio,
            data_leitura__date__lte=data_fim
        ).order_by('data_leitura')

        for i in range(1, len(leituras)):
            leitura_atual = leituras[i]
            leitura_anterior = leituras[i - 1]
            diferenca = float(leitura_atual.leitura - leitura_anterior.leitura)
            if diferenca <= 0:
                continue

            consumo_litros = diferenca * 1000
            consumo_total_periodo += consumo_litros

            dia = leitura_atual.data_leitura.date()
            consumo_por_dia[dia] = consumo_por_dia.get(dia, 0.0) + consumo_litros

            mes_key = (leitura_atual.data_leitura.year, leitura_atual.data_leitura.month)
            consumo_por_mes[mes_key] = consumo_por_mes.get(mes_key, 0.0) + consumo_litros

    datas_periodo = []
    dia_cursor = data_inicio
    while dia_cursor <= data_fim:
        datas_periodo.append(dia_cursor)
        consumo_por_dia.setdefault(dia_cursor, 0.0)
        dia_cursor += timedelta(days=1)

    # Sempre exibir todos os 12 meses do ano atual
    ano_atual = hoje.year
    meses_periodo = [(ano_atual, mes) for mes in range(1, 13)]
    # Garantir que todos os meses tenham valor 0 se não houver consumo
    for mes in range(1, 13):
        consumo_por_mes.setdefault((ano_atual, mes), 0.0)
    
    
    # Criar PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                          rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo do título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#7f8c8d'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Título
    elements.append(Paragraph(f"Relatório de Consumo - Lote {lote.numero} ({periodo_label})", title_style))
    elements.append(Paragraph(
        f"Tipo: {lote.get_tipo_display()} | Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')} | Gerado em: {agora.strftime('%d/%m/%Y %H:%M')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumo Geral
    elements.append(Paragraph("📊 Resumo Geral", heading_style))
    
    resumo_data = [
        ['Indicador', 'Valor'],
        ['Lote', lote.numero],
        ['Tipo', lote.get_tipo_display()],
        ['Período', periodo_label],
        ['Consumo Total no Período', f'{consumo_total_periodo:,.0f} L'],
        ['Hidrometros Ativos', str(hidrometros.count())],
    ]
    
    resumo_table = Table(resumo_data, colWidths=[3*inch, 2*inch])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Consumo Mensal
    elements.append(Paragraph("📅 Consumo Mensal", heading_style))
    
    mensal_data = [['Mês', 'Consumo (L)']]
    for (ano, mes) in meses_periodo:
        mes_nome = f'{nomes_meses[mes - 1]}/{str(ano)[-2:]}'
        mensal_data.append([mes_nome, f'{consumo_por_mes.get((ano, mes), 0.0):,.2f}'])
    
    mensal_table = Table(mensal_data, colWidths=[2*inch, 2*inch])
    mensal_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(mensal_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Gráfico de Consumo Mensal
    plt.figure(figsize=(10, 5))
    meses_labels = [f'{nomes_meses[mes - 1]}/{str(ano)[-2:]}' for (ano, mes) in meses_periodo]
    valores_mensais = [consumo_por_mes.get((ano, mes), 0.0) for (ano, mes) in meses_periodo]
    plt.bar(meses_labels, valores_mensais, color='#27ae60', alpha=0.7)
    plt.title(f'Consumo Mensal - Lote {lote.numero} (Litros)', fontsize=14, fontweight='bold')
    plt.xlabel('Mês', fontsize=11)
    plt.ylabel('Consumo (L)', fontsize=11)
    plt.xticks(rotation=45, ha='right')
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    
    # Salvar gráfico em buffer
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    # Adicionar imagem ao PDF
    img = Image(img_buffer, width=7*inch, height=3.5*inch)
    elements.append(img)
    elements.append(Spacer(1, 0.3*inch))

    leituras_periodo = Leitura.objects.filter(
        hidrometro__lote=lote,
        data_leitura__date__gte=data_inicio,
        data_leitura__date__lte=data_fim
    ).select_related('hidrometro').order_by('data_leitura')

    elements.append(PageBreak())
    elements.append(Paragraph("📋 Leituras no Período", heading_style))

    leituras_data = [[
        'Data/Hora',
        'Hidrômetro',
        'Leitura (m³)',
        'Consumo (L)',
        'Responsável',
        'Observações'
    ]]

    for leitura in leituras_periodo:
        consumo_litros = leitura.consumo_desde_ultima_leitura_litros()
        responsavel = leitura.responsavel or 'N/A'
        observacoes = leitura.observacoes or '—'
        if len(observacoes) > 60:
            observacoes = f"{observacoes[:57]}..."
        leituras_data.append([
            leitura.data_leitura.strftime('%d/%m/%Y %H:%M'),
            leitura.hidrometro.numero,
            f"{leitura.leitura}",
            f"{consumo_litros:,.0f}",
            responsavel,
            observacoes,
        ])

    leituras_table = Table(
        leituras_data,
        colWidths=[1.4*inch, 1.1*inch, 1.1*inch, 1.1*inch, 1.2*inch, 2.1*inch]
    )
    leituras_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))

    elements.append(leituras_table)
    elements.append(Spacer(1, 0.3*inch))

    leituras_com_foto = [leitura for leitura in leituras_periodo if leitura.foto]
    if leituras_com_foto:
        elements.append(PageBreak())
        elements.append(Paragraph("📷 Fotos das Leituras", heading_style))
        for leitura in leituras_com_foto:
            foto_path = getattr(leitura.foto, 'path', '')
            if not foto_path or not os.path.exists(foto_path):
                continue
            legenda = (
                f"Hidrômetro {leitura.hidrometro.numero} - "
                f"{leitura.data_leitura.strftime('%d/%m/%Y %H:%M')}"
            )
            elements.append(Paragraph(legenda, styles['Normal']))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Image(foto_path, width=6.5*inch, height=3.8*inch))
            elements.append(Spacer(1, 0.2*inch))
    
    
    # Construir PDF
    doc.build(
        elements,
        onFirstPage=_desenhar_marca_dagua_logo,
        onLaterPages=_desenhar_marca_dagua_logo
    )
    
    # Preparar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="relatorio_lote_{lote.numero}_{data_inicio.strftime("%Y%m%d")}_'
        f'{data_fim.strftime("%Y%m%d")}.pdf"'
    )
    
    return response


def exportar_graficos_lote_excel(request, lote_id):
    """Exporta os gráficos de consumo de um lote específico em Excel com gráficos"""
    import os
    os.environ.setdefault('MPLCONFIGDIR', '/tmp/matplotlib')
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    
    lote = get_object_or_404(Lote, id=lote_id)
    
    # Obter dados do lote (mesma lógica da view graficos_lote)
    agora = timezone.localtime(timezone.now())
    hoje = agora.date()
    periodo = request.GET.get('periodo', '30dias')
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')
    data_fim = hoje
    periodo_label = ''

    if periodo == '7dias':
        data_inicio = hoje - timedelta(days=7)
        periodo_label = 'Últimos 7 dias'
    elif periodo == '15dias':
        data_inicio = hoje - timedelta(days=15)
        periodo_label = 'Últimos 15 dias'
    elif periodo == '30dias':
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    elif periodo == 'mes_atual':
        data_inicio = hoje.replace(day=1)
        mes_nome = MESES_PT_BR[hoje.month]
        periodo_label = f'{mes_nome} de {hoje.year}'
    elif periodo == 'ano_atual':
        data_inicio = hoje.replace(month=1, day=1)
        periodo_label = f'Ano de {hoje.year}'
    elif periodo == 'personalizado' and data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            if data_fim > hoje:
                data_fim = hoje
            periodo_label = f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
        except (ValueError, TypeError):
            data_inicio = hoje - timedelta(days=30)
            data_fim = hoje
            periodo_label = 'Últimos 30 dias'
            periodo = '30dias'
    else:
        data_inicio = hoje - timedelta(days=30)
        periodo_label = 'Últimos 30 dias'
    
    hidrometros = lote.hidrometros.filter(ativo=True)
    
    if not hidrometros.exists():
        return HttpResponse("Nenhum hidrômetro ativo encontrado para este lote.", status=404)
    
    # Calcular consumo no periodo
    nomes_meses = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]
    consumo_total_periodo = 0.0
    consumo_por_dia = {}
    consumo_por_mes = {}

    for hidrometro in hidrometros:
        leituras = hidrometro.leituras.filter(
            data_leitura__date__gte=data_inicio,
            data_leitura__date__lte=data_fim
        ).order_by('data_leitura')

        for i in range(1, len(leituras)):
            leitura_atual = leituras[i]
            leitura_anterior = leituras[i - 1]
            diferenca = float(leitura_atual.leitura - leitura_anterior.leitura)
            if diferenca <= 0:
                continue

            consumo_litros = diferenca * 1000
            consumo_total_periodo += consumo_litros

            dia = leitura_atual.data_leitura.date()
            consumo_por_dia[dia] = consumo_por_dia.get(dia, 0.0) + consumo_litros

            mes_key = (leitura_atual.data_leitura.year, leitura_atual.data_leitura.month)
            consumo_por_mes[mes_key] = consumo_por_mes.get(mes_key, 0.0) + consumo_litros

    datas_periodo = []
    dia_cursor = data_inicio
    while dia_cursor <= data_fim:
        datas_periodo.append(dia_cursor)
        consumo_por_dia.setdefault(dia_cursor, 0.0)
        dia_cursor += timedelta(days=1)

    # Sempre exibir todos os 12 meses do ano atual
    ano_atual = hoje.year
    meses_periodo = [(ano_atual, mes) for mes in range(1, 13)]
    # Garantir que todos os meses tenham valor 0 se não houver consumo
    for mes in range(1, 13):
        consumo_por_mes.setdefault((ano_atual, mes), 0.0)
    
    
    # Criar Excel
    wb = Workbook()
    
    # Aba: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Título
    ws_resumo['A1'] = f'Relatório de Consumo - Lote {lote.numero} ({periodo_label})'
    ws_resumo['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_resumo['A1'].fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    ws_resumo['A1'].alignment = Alignment(horizontal='center')
    ws_resumo.merge_cells('A1:C1')
    
    ws_resumo['A2'] = (
        f'Tipo: {lote.get_tipo_display()} | Período: {data_inicio.strftime("%d/%m/%Y")} '
        f'a {data_fim.strftime("%d/%m/%Y")} | Gerado em: {agora.strftime("%d/%m/%Y %H:%M")}'
    )
    ws_resumo['A2'].alignment = Alignment(horizontal='center')
    ws_resumo.merge_cells('A2:C2')
    
    # Dados resumo
    ws_resumo['A4'] = 'Indicador'
    ws_resumo['B4'] = 'Valor'
    ws_resumo['A4'].font = Font(bold=True)
    ws_resumo['B4'].font = Font(bold=True)
    
    resumo_dados = [
        ['Lote', lote.numero],
        ['Tipo', lote.get_tipo_display()],
        ['Período', periodo_label],
        ['Consumo Total no Período', f'{consumo_total_periodo:,.0f} L'],
        ['Hidrometros Ativos', hidrometros.count()],
    ]
    
    for idx, (indicador, valor) in enumerate(resumo_dados, start=5):
        ws_resumo[f'A{idx}'] = indicador
        ws_resumo[f'B{idx}'] = valor
    
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 20
    
    # Aba: Consumo Mensal
    ws_mensal = wb.create_sheet("Consumo Mensal")
    
    ws_mensal['A1'] = 'Mês'
    ws_mensal['B1'] = 'Consumo (L)'
    ws_mensal['A1'].font = Font(bold=True)
    ws_mensal['B1'].font = Font(bold=True)
    
    for idx, (ano, mes) in enumerate(meses_periodo, start=2):
        ws_mensal[f'A{idx}'] = f'{nomes_meses[mes - 1]}/{str(ano)[-2:]}'
        ws_mensal[f'B{idx}'] = round(consumo_por_mes.get((ano, mes), 0.0), 2)
    
    # Gerar gráfico com matplotlib (igual ao PDF)
    plt.figure(figsize=(12, 6))
    meses_labels = [f'{nomes_meses[mes - 1]}/{str(ano)[-2:]}' for (ano, mes) in meses_periodo]
    valores_mensais = [consumo_por_mes.get((ano, mes), 0.0) for (ano, mes) in meses_periodo]
    plt.bar(meses_labels, valores_mensais, color='#27ae60', alpha=0.7)
    plt.title(f'Consumo Mensal - Lote {lote.numero} (Litros)', fontsize=14, fontweight='bold')
    plt.xlabel('Mês', fontsize=11)
    plt.ylabel('Consumo (L)', fontsize=11)
    plt.xticks(rotation=45, ha='right')
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    
    # Salvar gráfico em buffer
    img_buffer_mensal = io.BytesIO()
    plt.savefig(img_buffer_mensal, format='png', dpi=100, bbox_inches='tight')
    img_buffer_mensal.seek(0)
    plt.close()
    
    # Adicionar imagem ao Excel
    from openpyxl.drawing.image import Image as XLImage
    img_mensal = XLImage(img_buffer_mensal)
    img_mensal.width = 600
    img_mensal.height = 300
    ws_mensal.add_image(img_mensal, "D2")
    
    ws_mensal.column_dimensions['A'].width = 15
    ws_mensal.column_dimensions['B'].width = 15
    
    # Aba: Consumo Diário
    ws_diario_lote = wb.create_sheet("Consumo Diário")
    
    ws_diario_lote['A1'] = 'Dia'
    ws_diario_lote['B1'] = 'Consumo (L)'
    ws_diario_lote['A1'].font = Font(bold=True)
    ws_diario_lote['B1'].font = Font(bold=True)
    
    for idx, dia in enumerate(datas_periodo, start=2):
        ws_diario_lote[f'A{idx}'] = dia.strftime('%d/%m/%Y')
        ws_diario_lote[f'B{idx}'] = round(consumo_por_dia.get(dia, 0.0), 2)
    
    # Gerar gráfico com matplotlib (igual ao PDF)
    plt.figure(figsize=(12, 6))
    dias_labels = [d.strftime('%d/%m') for d in datas_periodo]
    valores_diarios_lote = [consumo_por_dia.get(d, 0.0) for d in datas_periodo]
    plt.plot(dias_labels, valores_diarios_lote, marker='o', color='#3498db', linewidth=2, markersize=4)
    plt.title(f'Consumo Diário - Lote {lote.numero} ({periodo_label})', fontsize=14, fontweight='bold')
    plt.xlabel('Dia', fontsize=11)
    plt.ylabel('Consumo (L)', fontsize=11)
    plt.xticks(rotation=45, ha='right')
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    
    # Salvar gráfico em buffer
    img_buffer_diario_lote = io.BytesIO()
    plt.savefig(img_buffer_diario_lote, format='png', dpi=100, bbox_inches='tight')
    img_buffer_diario_lote.seek(0)
    plt.close()
    
    # Adicionar imagem ao Excel
    img_diario_lote = XLImage(img_buffer_diario_lote)
    img_diario_lote.width = 600
    img_diario_lote.height = 300
    ws_diario_lote.add_image(img_diario_lote, "D2")
    
    ws_diario_lote.column_dimensions['A'].width = 15
    ws_diario_lote.column_dimensions['B'].width = 15

    leituras_periodo = Leitura.objects.filter(
        hidrometro__lote=lote,
        data_leitura__date__gte=data_inicio,
        data_leitura__date__lte=data_fim
    ).select_related('hidrometro').order_by('data_leitura')

    ws_leituras = wb.create_sheet("Leituras")
    ws_leituras['A1'] = 'Data/Hora'
    ws_leituras['B1'] = 'Hidrômetro'
    ws_leituras['C1'] = 'Leitura (m³)'
    ws_leituras['D1'] = 'Consumo (L)'
    ws_leituras['E1'] = 'Responsável'
    ws_leituras['F1'] = 'Observações'
    ws_leituras['G1'] = 'Foto'

    for col in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']:
        ws_leituras[col].font = Font(bold=True)

    for idx, leitura in enumerate(leituras_periodo, start=2):
        consumo_litros = leitura.consumo_desde_ultima_leitura_litros()
        ws_leituras[f'A{idx}'] = leitura.data_leitura.strftime('%d/%m/%Y %H:%M')
        ws_leituras[f'B{idx}'] = leitura.hidrometro.numero
        ws_leituras[f'C{idx}'] = float(leitura.leitura)
        ws_leituras[f'D{idx}'] = round(consumo_litros, 2)
        ws_leituras[f'E{idx}'] = leitura.responsavel or 'N/A'
        ws_leituras[f'F{idx}'] = leitura.observacoes or '—'

        if leitura.foto:
            foto_path = getattr(leitura.foto, 'path', '')
            if foto_path and os.path.exists(foto_path):
                img = XLImage(foto_path)
                img.width = 120
                img.height = 90
                ws_leituras.add_image(img, f'G{idx}')
                ws_leituras.row_dimensions[idx].height = 70

    ws_leituras.column_dimensions['A'].width = 18
    ws_leituras.column_dimensions['B'].width = 15
    ws_leituras.column_dimensions['C'].width = 14
    ws_leituras.column_dimensions['D'].width = 14
    ws_leituras.column_dimensions['E'].width = 18
    ws_leituras.column_dimensions['F'].width = 40
    ws_leituras.column_dimensions['G'].width = 22
    
    # Salvar e retornar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="relatorio_lote_{lote.numero}_{data_inicio.strftime("%Y%m%d")}_'
        f'{data_fim.strftime("%Y%m%d")}.xlsx"'
    )
    
    return response

